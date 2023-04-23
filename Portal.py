import pandas as pd
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver as opcoesSelenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from time import sleep

NomeCaminhoArquivo = "Consulta_SDPWIN.xlsx"
PlanilhaAberta= load_workbook(filename=NomeCaminhoArquivo)

Serviço = Service(ChromeDriverManager().install())




workbook = openpyxl.load_workbookworkbook = openpyxl.load_workbook('Consulta_SDPWIN.xlsx')

# Selecionando a planilha e células que contêm as informações desejadas
worksheet = workbook['Limpa']
informacao1 = worksheet['A1'].value
informacao2 = worksheet['C1'].value

# Iniciando o driver do navegador
driver = webdriver.Chrome()

# Acessando o formulário desejado
driver.get("https://cpag.prf.gov.br/nada_consta/index.jsf")

# Preenchendo o formulário com as informações desejadas
campo1 = driver.find_element_by_id('formConsultarExterno:placa').send_keyssend_keys(informacao1)

campo2 = driver.find_element_by_id('formConsultarExterno:renavam').send_keys.send_keys(informacao2)


sleep(10)